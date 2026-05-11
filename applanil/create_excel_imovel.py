from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import (ColorScaleRule, DataBarRule,
                                       FormulaRule, CellIsRule)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── PALETA ────────────────────────────────────────────────────────────────────
AZUL_ESCURO  = "FF1A3C6E"
AZUL_MED     = "FF2E75B6"
AZUL_VIVO    = "FF3A86FF"
AZUL_LIGHT   = "FFD6E8FF"
CORAL        = "FFFF6B6B"
CORAL_LIGHT  = "FFFFD6D6"
VERDE        = "FF06D6A0"
VERDE_DARK   = "FF047857"
VERDE_LIGHT  = "FFD0F5EC"
AMARELO      = "FFFFBE0B"
AMAR_LIGHT   = "FFFFF3CC"
ROXO         = "FF8338EC"
ROXO_LIGHT   = "FFEDE7FF"
LARANJA      = "FFFF6D00"
LARAN_LIGHT  = "FFFFE5CC"
CINZA_DARK   = "FF2D2D2D"
CINZA_MED    = "FF6B6B6B"
CINZA_CLARO  = "FFF0F0F0"
BRANCO       = "FFFFFFFF"
ROSA         = "FFFF70A6"
ROSA_LIGHT   = "FFFFEEF6"

def fill(c): return PatternFill("solid", fgColor=c)
def fnt(bold=False, size=10, color="FF000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")
def aln(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def brd(color="FFD0D0D0", style="thin"):
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def brd_thick(color="FF888888"):
    t = Side(border_style="medium", color=color)
    return Border(left=t, right=t, top=t, bottom=t)

def sc(ws, ref, val=None, bold=False, size=10, color=CINZA_DARK,
       bg=BRANCO, h="center", wrap=False, fmt=None, italic=False, border_c="FFD0D0D0"):
    c = ws[ref]
    if val is not None: c.value = val
    c.font = fnt(bold, size, color, italic)
    c.fill = fill(bg)
    c.alignment = aln(h, "center", wrap)
    c.border = brd(border_c)
    if fmt: c.number_format = fmt
    return c

def input_sc(ws, ref, val=None, fmt=None):
    c = ws[ref]
    if val is not None: c.value = val
    c.font = Font(bold=False, size=10, color="FF003580", name="Calibri")
    c.fill = fill(AMAR_LIGHT)
    c.alignment = aln("center", "center")
    c.border = Border(
        left=Side("medium", color="FFFF6D00"),
        right=Side("medium", color="FFFF6D00"),
        top=Side("thin", color="FFAAAAAA"),
        bottom=Side("thin", color="FFAAAAAA"),
    )
    if fmt: c.number_format = fmt
    return c

def calc_sc(ws, ref, formula, fmt=None, bg=AZUL_LIGHT):
    c = ws[ref]
    c.value = formula
    c.font = fnt(False, 10, AZUL_ESCURO)
    c.fill = fill(bg)
    c.alignment = aln("center", "center")
    c.border = brd("FF3A86FF")
    if fmt: c.number_format = fmt
    return c

def total_sc(ws, ref, formula, fmt=None, bg=VERDE, color=BRANCO):
    c = ws[ref]
    c.value = formula
    c.font = fnt(True, 10, color)
    c.fill = fill(bg)
    c.alignment = aln("center", "center")
    c.border = brd(VERDE_DARK, "medium")
    if fmt: c.number_format = fmt
    return c

# ════════════════════════════════════════════════════════════════════════════════
# ABA ÚNICA — FINANCIAMENTO
# ════════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "🏠 Financiamento"
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = "3A86FF"

# Larguras das colunas
# A=decor, B=Nº, C=Mês/Ano, D=Data Pgto, E=Sistema, F=Saldo Devedor,
# G=Parcela Total, H=Amortização, I=Juros, J=Amort.Adicional,
# K=Seguro, L=Data Pagamento, M=Valor Pago, N=Situação
col_widths = {
    "A": 2.5, "B": 5,  "C": 11, "D": 11, "E": 8,
    "F": 17,  "G": 14, "H": 14, "I": 13, "J": 14,
    "K": 11,  "L": 11, "M": 13, "N": 13, "O": 14
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# Altura padrão linhas
for r in range(1, 450):
    ws.row_dimensions[r].height = 18

# Decor col A
for r in range(1, 450):
    c = ws[f"A{r}"]
    c.fill = fill(AZUL_VIVO if r % 2 == 1 else AZUL_MED)

# ── ROW 1: Banner principal ────────────────────────────────────────────────────
ws.row_dimensions[1].height = 50
for col in "BCDEFGHIJKLMN":
    ws[f"{col}1"].fill = fill(AZUL_ESCURO)
    ws[f"{col}1"].border = brd(AZUL_ESCURO)
c = ws["B1"]
c.value = "🏠  TABELA DE AMORTIZAÇÃO DO FINANCIAMENTO IMOBILIÁRIO"
c.font = Font(bold=True, size=18, color=BRANCO, name="Calibri")
c.fill = fill(AZUL_ESCURO)
c.alignment = aln("center", "center")
c.border = brd(AZUL_ESCURO)
ws.merge_cells("B1:N1")

# ── ROW 2: Config inputs ──────────────────────────────────────────────────────
ws.row_dimensions[2].height = 8

# ── BLOCO CONFIGURAÇÃO (linhas 3-10) ─────────────────────────────────────────
ws.row_dimensions[3].height = 26
c = ws["B3"]
c.value = "⚙️  CONFIGURAÇÕES DO FINANCIAMENTO"
c.font = fnt(True, 12, BRANCO)
c.fill = fill(ROXO)
c.alignment = aln("left", "center")
c.border = brd_thick(ROXO)
ws.merge_cells("B3:N3")

# Linha de config — esquerda
cfg_left = [
    ("B4", "💰 Valor Financiado (R$):"),
    ("B5", "📅 Mês/Ano de Início:"),
    ("B6", "⏳ Prazo (meses):"),
    ("B7", "📈 Taxa de Juros (% a.m.):"),
]
for ref, lbl in cfg_left:
    ws[ref].value = lbl
    ws[ref].font = Font(bold=True, size=9, color=BRANCO, name="Calibri")
    ws[ref].fill = fill(ROXO)
    ws[ref].alignment = aln("right", "center")
    ws[ref].border = brd(ROXO)

ws.merge_cells("B4:C4"); ws.merge_cells("B5:C5")
ws.merge_cells("B6:C6"); ws.merge_cells("B7:C7")

# Inputs config esquerda
input_sc(ws, "D4", fmt='R$ #,##0.00')
input_sc(ws, "D5", "Jan/2025")
input_sc(ws, "D6", 360, fmt='0')
input_sc(ws, "D7", fmt='0.000%')

# Labels config direita
cfg_right = [
    ("F4", "🏦 Sistema:"),
    ("F5", "🛡️ Seguro Mensal (R$):"),
    ("F6", "📊 Amortização Adicional Padrão:"),
    ("F7", ""),
]
for ref, lbl in cfg_right:
    ws[ref].value = lbl
    ws[ref].font = Font(bold=True, size=9, color=BRANCO, name="Calibri")
    ws[ref].fill = fill(ROXO)
    ws[ref].alignment = aln("right", "center")

ws.merge_cells("F4:G4"); ws.merge_cells("F5:G5"); ws.merge_cells("F6:G6")

# Dropdown sistema
dv_sistema = DataValidation(type="list", formula1='"SAC,PRICE"', allow_blank=False)
ws.add_data_validation(dv_sistema)
input_sc(ws, "H4", "SAC")
dv_sistema.add(ws["H4"])

input_sc(ws, "H5", fmt='R$ #,##0.00')
input_sc(ws, "H6", fmt='R$ #,##0.00')

# Config cols I-N linha 4-7
for row in range(4, 8):
    for col in "IJKLMN":
        ws[f"{col}{row}"].fill = fill(ROXO_LIGHT)
        ws[f"{col}{row}"].border = brd("FFCCBBEE")

# Cells de info calculada
sc(ws, "J4", "📋 Parcela PRICE:", bold=True, size=9, color=ROXO, bg=ROXO_LIGHT, h="right")
ws.merge_cells("J4:K4")
input_sc(ws, "L4", fmt='R$ #,##0.00')
ws.merge_cells("L4:N4")

sc(ws, "J5", "📋 Parcela SAC (referência):", bold=True, size=9, color=ROXO, bg=ROXO_LIGHT, h="right")
ws.merge_cells("J5:K5")
sc(ws, "L5", "(*continua variável)", bold=False, size=8, color="FF999999", bg=ROXO_LIGHT, h="center")
ws.merge_cells("L5:N5")

sc(ws, "J6", "💰 Total a Pagar:", bold=True, size=9, color=ROXO, bg=ROXO_LIGHT, h="right")
ws.merge_cells("J6:K6")
c = ws["L6"]
c.value = '=IFERROR(IF(H4="PRICE",-PMT(D7,D6,D4)*D6,((D4/D6)+D4*D7+(D4/D6+0)*D6)/2),"—")'
c.font = fnt(True, 10, ROXO)
c.fill = fill(ROXO_LIGHT)
c.alignment = aln("center", "center")
c.border = brd(ROXO, "medium")
c.number_format = 'R$ #,##0.00'
ws.merge_cells("L6:N6")

sc(ws, "J7", "💸 Total de Juros:", bold=True, size=9, color=ROXO, bg=ROXO_LIGHT, h="right")
ws.merge_cells("J7:K7")
c = ws["L7"]
c.value = '=IFERROR(L6-D4,"—")'
c.font = fnt(True, 10, ROXO)
c.fill = fill(ROXO_LIGHT)
c.alignment = aln("center", "center")
c.border = brd(ROXO, "medium")
c.number_format = 'R$ #,##0.00'
ws.merge_cells("L7:N7")

# Separador
ws.row_dimensions[8].height = 6
ws.row_dimensions[9].height = 6

# ── TOTALIZADORES ─────────────────────────────────────────────────────────────
ws.row_dimensions[10].height = 28
tot_labels = [
    ("B", "✅ PARCELAS PAGAS"),
    ("D", "💰 TOTAL AMORTIZADO"),
    ("F", "💸 TOTAL JUROS PAGOS"),
    ("H", "🛡️ TOTAL SEGUROS"),
    ("J", "📊 SALDO DEVEDOR ATUAL"),
    ("L", "📈 % QUITADO"),
]
tot_colors = [VERDE, AZUL_VIVO, CORAL, AMARELO, ROXO, LARANJA]
tot_color_txt = [BRANCO, BRANCO, BRANCO, CINZA_DARK, BRANCO, BRANCO]
for (col, lbl), bg, fg in zip(tot_labels, tot_colors, tot_color_txt):
    c = ws[f"{col}10"]
    c.value = lbl
    c.font = Font(bold=True, size=8, color=fg, name="Calibri")
    c.fill = fill(bg)
    c.alignment = aln("center", "center", wrap=True)
    c.border = brd_thick()
    next_col = get_column_letter(ord(col)-64+1)
    ws.merge_cells(f"{col}10:{next_col}10")

ws.row_dimensions[11].height = 28
# Fórmulas dos totalizadores (linhas 13..412 são os dados)
formulas = [
    ("B11", '=COUNTIF(N13:N412,"✅ PAGO")', None, VERDE, BRANCO),
    ("D11", "=IFERROR(SUM(H13:H412)+SUM(J13:J412),0)", 'R$ #,##0.00', AZUL_VIVO, BRANCO),
    ("F11", "=IFERROR(SUM(I13:I412),0)", 'R$ #,##0.00', CORAL, BRANCO),
    ("H11", "=IFERROR(SUM(K13:K412),0)", 'R$ #,##0.00', AMARELO, CINZA_DARK),
    ("J11", "=IFERROR(F13,0)", 'R$ #,##0.00', ROXO, BRANCO),  # saldo mais recente pendente
    ("L11", '=IFERROR(1-F13/D4,"—")', '0.0%', LARANJA, BRANCO),
]
for ref, formula, fmt, bg, fg in formulas:
    c = ws[ref]
    c.value = formula
    c.font = Font(bold=True, size=12, color=fg, name="Calibri")
    c.fill = fill(bg)
    c.alignment = aln("center", "center")
    c.border = brd_thick()
    if fmt: c.number_format = fmt
    next_col = get_column_letter(ord(ref[0])-64+1)
    ws.merge_cells(f"{ref}:{next_col}11")

# Separador
ws.row_dimensions[12].height = 8

# ── HEADER DA TABELA ──────────────────────────────────────────────────────────
ws.row_dimensions[13].height = 30
headers = [
    ("B", "#"),
    ("C", "📅 MÊS / ANO"),
    ("D", "📆 VENCIMENTO"),
    ("E", "🏦 SIST."),
    ("F", "💰 SALDO DEVEDOR"),
    ("G", "💳 PARCELA (de L4/L5)"),
    ("H", "🔨 AMORTIZAÇÃO"),
    ("I", "📈 JUROS"),
    ("J", "➕ AMORT. ADICIONAL"),
    ("K", "🛡️ SEGURO"),
    ("L", "📅 DATA PGTO"),
    ("M", "✏️ VALOR PAGO"),
    ("N", "📊 TOTAL ACUMULADO"),
    ("O", "🚦 SITUAÇÃO"),
]
HEADER_BG = AZUL_ESCURO
for col, lbl in headers:
    c = ws[f"{col}13"]
    c.value = lbl
    c.font = Font(bold=True, size=8, color=BRANCO, name="Calibri")
    c.fill = fill(HEADER_BG)
    c.alignment = aln("center", "center", wrap=True)
    c.border = brd(AZUL_VIVO, "medium")

# ── LINHAS DE DADOS (14 a 373 = 360 meses) ───────────────────────────────────
N_MESES = 360
START_ROW = 14

for i in range(N_MESES):
    r = START_ROW + i
    ws.row_dimensions[r].height = 19
    alt = i % 2 == 0

    bg_base   = "FFF0F7FF" if alt else BRANCO
    bg_saldo  = "FFE8F4FF" if alt else "FFF5FAFF"
    bg_juros  = CORAL_LIGHT if alt else "FFFFE8E8"
    bg_amort  = VERDE_LIGHT if alt else "FFE0F9F2"
    bg_seg    = LARAN_LIGHT if alt else "FFFFF0E0"

    # B — número da parcela
    c = ws[f"B{r}"]
    c.value = i + 1
    c.font = Font(bold=True, size=8, color=BRANCO, name="Calibri")
    c.fill = fill(AZUL_MED if alt else AZUL_VIVO)
    c.alignment = aln("center", "center")
    c.border = brd(AZUL_ESCURO)

    # C — Mês/Ano (calculado)
    c = ws[f"C{r}"]
    if r == START_ROW:
        c.value = "=D5"
    else:
        # incrementa mês baseado no texto do anterior
        c.value = f'=TEXT(DATE(YEAR(DATE(1,MID(C{r-1},5,4),1))+IF(MID(C{r-1},1,3)="Dez",1,0),IF(MID(C{r-1},1,3)="Dez",1,MONTH(DATE(1,MID(C{r-1},1,3),1))+1),1),"mmm/yyyy")'
    # Simpler approach — just reference by offset
    c.value = f'=IFERROR(TEXT(EDATE(DATE(VALUE(RIGHT(D5,4)),MATCH(LEFT(D5,3),{{"Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"}},0),1),{i}),"mmm/yyyy")'
    c.font = fnt(False, 9, AZUL_ESCURO)
    c.fill = fill(bg_base)
    c.alignment = aln("center", "center")
    c.border = brd("FFB0CCEE")

    # D — Data vencimento (dia 5 de cada mês)
    c = ws[f"D{r}"]
    c.value = f'=IFERROR(EDATE(DATE(VALUE(RIGHT(D5,4)),MATCH(LEFT(D5,3),{{"Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"}},0),5),{i}),"—")'
    c.font = fnt(False, 9, AZUL_ESCURO)
    c.fill = fill(bg_base)
    c.alignment = aln("center", "center")
    c.border = brd("FFB0CCEE")
    c.number_format = "DD/MM/YYYY"

    # E — Sistema
    c = ws[f"E{r}"]
    c.value = "=$H$4"
    c.font = fnt(True, 8, ROXO)
    c.fill = fill(ROXO_LIGHT)
    c.alignment = aln("center", "center")
    c.border = brd("FFCCBBEE")

    # F — Saldo Devedor
    c = ws[f"F{r}"]
    if r == START_ROW:
        c.value = "=D4"
    else:
        c.value = f"=IFERROR(F{r-1}-H{r-1}-J{r-1},0)"
    c.font = fnt(True, 9, AZUL_ESCURO)
    c.fill = fill(bg_saldo)
    c.alignment = aln("center", "center")
    c.border = brd("FF3A86FF", "medium")
    c.number_format = 'R$ #,##0.00'

    # G — Parcela Total (PRICE constante de L4, SAC diminui com juros)
    c = ws[f"G{r}"]
    c.value = f'=IF($H$4="PRICE",$L$4,IF($H$4="SAC",($D$4/$D$6)+I{r}+K{r},0))'
    c.font = fnt(True, 9, CINZA_DARK)
    c.fill = fill(bg_base)
    c.alignment = aln("center", "center")
    c.border = brd("FF888888")
    c.number_format = 'R$ #,##0.00'

    # H — Amortização
    c = ws[f"H{r}"]
    c.value = f'=IFERROR(IF($H$4="SAC",$D$4/$D$6,G{r}-I{r}-K{r}),0)'
    c.font = fnt(False, 9, VERDE_DARK)
    c.fill = fill(bg_amort)
    c.alignment = aln("center", "center")
    c.border = brd(VERDE)
    c.number_format = 'R$ #,##0.00'

    # I — Juros
    c = ws[f"I{r}"]
    c.value = f"=IFERROR(F{r}*$D$7,0)"
    c.font = fnt(False, 9, "FF8B0000")
    c.fill = fill(bg_juros)
    c.alignment = aln("center", "center")
    c.border = brd(CORAL)
    c.number_format = 'R$ #,##0.00'

    # J — Amortização Adicional (input)
    c = ws[f"J{r}"]
    c.font = Font(bold=False, size=9, color="FF003580", name="Calibri")
    c.fill = fill(AMAR_LIGHT)
    c.alignment = aln("center", "center")
    c.border = Border(
        left=Side("thin", color="FFFF6D00"),
        right=Side("thin", color="FFFF6D00"),
        top=Side("thin", color="FFCCCCCC"),
        bottom=Side("thin", color="FFCCCCCC"),
    )
    c.number_format = 'R$ #,##0.00'

    # K — Seguro
    c = ws[f"K{r}"]
    c.value = "=IFERROR($H$5,0)"
    c.font = fnt(False, 9, "FF5C3000")
    c.fill = fill(bg_seg)
    c.alignment = aln("center", "center")
    c.border = brd(LARANJA)
    c.number_format = 'R$ #,##0.00'

    # L — Data de Pagamento (input)
    c = ws[f"L{r}"]
    c.font = Font(bold=False, size=9, color="FF003580", name="Calibri")
    c.fill = fill(AMAR_LIGHT)
    c.alignment = aln("center", "center")
    c.border = Border(
        left=Side("thin", color="FFFF6D00"),
        right=Side("thin", color="FFFF6D00"),
        top=Side("thin", color="FFCCCCCC"),
        bottom=Side("thin", color="FFCCCCCC"),
    )
    c.number_format = "DD/MM/YYYY"

    # M — Valor Pago (input)
    c = ws[f"M{r}"]
    c.font = Font(bold=False, size=9, color="FF003580", name="Calibri")
    c.fill = fill(AMAR_LIGHT)
    c.alignment = aln("center", "center")
    c.border = Border(
        left=Side("thin", color="FFFF6D00"),
        right=Side("thin", color="FFFF6D00"),
        top=Side("thin", color="FFCCCCCC"),
        bottom=Side("thin", color="FFCCCCCC"),
    )
    c.number_format = 'R$ #,##0.00'

    # N — Total Acumulado (fórmula)
    c = ws[f"N{r}"]
    if i == 0:
        c.value = f"=M{r}"
    else:
        c.value = f"=IFERROR(N{r-1}+M{r},N{r-1})"
    c.font = fnt(False, 9, "FF006050")
    c.fill = fill("FFD5F5F3")
    c.alignment = aln("center", "center")
    c.border = brd("FF4ECDC4")
    c.number_format = 'R$ #,##0.00'

    # O — Situação (fórmula automática)
    c = ws[f"O{r}"]
    c.value = (
        f'=IF(M{r}="","⏳ PENDENTE",'
        f'IF(M{r}>=G{r},"✅ PAGO",'
        f'"⚠️ PARCIAL"))'
    )
    c.font = fnt(True, 9, CINZA_DARK)
    c.fill = fill(AMAR_LIGHT)
    c.alignment = aln("center", "center")
    c.border = brd("FF999999")

# ── FORMATAÇÃO CONDICIONAL — SITUAÇÃO ─────────────────────────────────────────
data_range = f"O{START_ROW}:O{START_ROW+N_MESES-1}"

ws.conditional_formatting.add(data_range,
    FormulaRule(
        formula=[f'O{START_ROW}="✅ PAGO"'],
        fill=PatternFill("solid", fgColor="FFC6EFCE"),
        font=Font(bold=True, color="FF1E7E34", name="Calibri")
    ))
ws.conditional_formatting.add(data_range,
    FormulaRule(
        formula=[f'O{START_ROW}="⚠️ PARCIAL"'],
        fill=PatternFill("solid", fgColor="FFFFD6D6"),
        font=Font(bold=True, color="FFCC0000", name="Calibri")
    ))
ws.conditional_formatting.add(data_range,
    FormulaRule(
        formula=[f'O{START_ROW}="⏳ PENDENTE"'],
        fill=PatternFill("solid", fgColor="FFFFEB9C"),
        font=Font(bold=True, color="FF9C6500", name="Calibri")
    ))

# ── FORMATAÇÃO CONDICIONAL — LINHAS INTEIRAS (pago/pendente) ──────────────────
full_range = f"B{START_ROW}:N{START_ROW+N_MESES-1}"
ws.conditional_formatting.add(full_range,
    FormulaRule(
        formula=[f'$N{START_ROW}="✅ PAGO"'],
        fill=PatternFill("solid", fgColor="FFE8FFF4"),
        font=Font(color="FF1E7E34", name="Calibri")
    ))

# ── DATA BAR — SALDO DEVEDOR (diminui visualmente) ───────────────────────────
saldo_range = f"F{START_ROW}:F{START_ROW+N_MESES-1}"
ws.conditional_formatting.add(saldo_range,
    DataBarRule(
        start_type="num", start_value=0,
        end_type="max",
        color="3A86FF",
        showValue=True,
        minLength=0, maxLength=100
    ))

# ── DATA BAR — JUROS ──────────────────────────────────────────────────────────
juros_range = f"I{START_ROW}:I{START_ROW+N_MESES-1}"
ws.conditional_formatting.add(juros_range,
    DataBarRule(
        start_type="num", start_value=0,
        end_type="max",
        color="FF6B6B",
        showValue=True,
        minLength=0, maxLength=100
    ))

# ── COLOR SCALE — AMORTIZAÇÃO (verde quando maior) ───────────────────────────
amort_range = f"H{START_ROW}:H{START_ROW+N_MESES-1}"
ws.conditional_formatting.add(amort_range,
    ColorScaleRule(
        start_type="min", start_color="FFE0FFEF",
        end_type="max", end_color="FF06D6A0"
    ))

# ── LINHA TOTAL FINAL ─────────────────────────────────────────────────────────
tot_r = START_ROW + N_MESES
ws.row_dimensions[tot_r].height = 26
sc(ws, f"B{tot_r}", "🏆 TOTAIS", bold=True, size=10, color=BRANCO, bg=VERDE_DARK, h="center")
ws.merge_cells(f"B{tot_r}:E{tot_r}")

totais = [
    (f"F{tot_r}", "=SUM(F14:F413)", 'R$ #,##0.00'),  # não faz sentido somar saldo, mas deixa
    (f"G{tot_r}", f"=SUM(G{START_ROW}:G{START_ROW+N_MESES-1})", 'R$ #,##0.00'),
    (f"H{tot_r}", f"=SUM(H{START_ROW}:H{START_ROW+N_MESES-1})", 'R$ #,##0.00'),
    (f"I{tot_r}", f"=SUM(I{START_ROW}:I{START_ROW+N_MESES-1})", 'R$ #,##0.00'),
    (f"J{tot_r}", f"=SUM(J{START_ROW}:J{START_ROW+N_MESES-1})", 'R$ #,##0.00'),
    (f"K{tot_r}", f"=SUM(K{START_ROW}:K{START_ROW+N_MESES-1})", 'R$ #,##0.00'),
    (f"M{tot_r}", f"=SUM(M{START_ROW}:M{START_ROW+N_MESES-1})", 'R$ #,##0.00'),
]
for ref, formula, fmt in totais:
    total_sc(ws, ref, formula, fmt, VERDE_DARK, BRANCO)

# F total = não somar saldo, mostrar "VEJA SALDO"
ws[f"F{tot_r}"].value = "SALDO ↑"
ws[f"F{tot_r}"].font = Font(bold=True, size=9, color=BRANCO, name="Calibri")
ws[f"F{tot_r}"].fill = fill(VERDE_DARK)
ws[f"F{tot_r}"].alignment = aln("center", "center")

# ── FREEZE PANES ─────────────────────────────────────────────────────────────
ws.freeze_panes = "B14"  # congela header e config

# ── LEGENDA ──────────────────────────────────────────────────────────────────
leg_r = tot_r + 2
ws.row_dimensions[leg_r].height = 22
leg = ws[f"B{leg_r}"]
leg.value = "🎨  LEGENDA:"
leg.font = fnt(True, 9, CINZA_DARK)
leg.fill = fill(CINZA_CLARO)
leg.alignment = aln("left", "center")
ws.merge_cells(f"B{leg_r}:N{leg_r}")

legendas = [
    (leg_r+1, AMAR_LIGHT, LARANJA, "✏️  Amarelo / laranja → Célula de entrada (você preenche aqui)"),
    (leg_r+2, VERDE_LIGHT, VERDE_DARK, "🟢  Verde claro → Amortização / valores calculados positivos"),
    (leg_r+3, CORAL_LIGHT, "FF8B0000", "🔴  Vermelho claro → Juros (diminui com SAC ao longo do tempo)"),
    (leg_r+4, AZUL_LIGHT, AZUL_ESCURO, "🔵  Azul claro → Saldo devedor (barra visual mostra quanto falta)"),
    (leg_r+5, "FFC6EFCE", "FF1E7E34", "✅  Verde escuro → Parcela PAGA"),
    (leg_r+6, "FFFFEB9C", "FF9C6500", "⏳  Amarelo → Parcela PENDENTE"),
    (leg_r+7, "FFFFD6D6", "FFCC0000", "⚠️   Vermelho → Parcela PARCIAL (valor pago menor que parcela)"),
]
for row, bg, fg, txt in legendas:
    ws.row_dimensions[row].height = 18
    c = ws[f"B{row}"]
    c.value = "   "
    c.fill = fill(bg)
    c.border = brd("FF999999")
    c2 = ws[f"C{row}"]
    c2.value = txt
    c2.font = fnt(False, 9, fg)
    c2.fill = fill(BRANCO)
    c2.alignment = aln("left", "center")
    c2.border = brd("FFD0D0D0")
    ws.merge_cells(f"C{row}:N{row}")

OUTPUT = "Imovel_Financiado_PERSONALIZADO.xlsx"
wb.save(OUTPUT)
print(f"✅ Salvo em {OUTPUT}")